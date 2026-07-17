from flask import Flask, render_template,request,session,redirect,send_file
import sqlite3
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "LR2026"
@app.route("/")
def inicio():

    if "usuario" not in session:

       return """
       <html>

       <head>
       <link rel="stylesheet" href="/static/estilo.css">
       </head>

       <body>

       <img src="/static/Logo LR.jpeg" width="220">

       <h1>Acceso restringido</h1>

       <p>Debes iniciar sesión para acceder al sistema.</p>

       <a href="/login">
       <button>🔑 Iniciar sesión</button>
       </a>

       </body>

       </html>
       """
    return render_template("index.html")
@app.route("/login", methods=["GET","POST"])
def login():

    if request.method == "POST":

        usuario = request.form["usuario"]
        password = request.form["password"]

        print("Usuario recibido:",repr(usuario))
        print("Password recibido:",repr(password))

        if usuario =="LR" and password =="LR2010":
            print("Usuario recibido:", usuario)
            print("Password recibido:", password)

            session["usuario"] = usuario

            return """
            <html>

            <head>
            <link rel="stylesheet" href="/static/estilo.css">
            </head>

            <body>

            <img src="/static/Logo LR.jpeg" class="logo">

            <h1>¡Bienvenido!</h1>

            <h2>Acceso correcto</h2>

            <p>Has iniciado sesión correctamente.</p>

            <a href="/">
            <button>🏠 Ir al menú principal</button>
            </a>

            </body>

            </html>
            """

        return """
        <h1>Usuario o contraseña incorrectos</h1>

        <a href='/login'>
            <button>Intentar de nuevo</button>
        </a>
        """

    return """
        <html>

        <head>
        <link rel="stylesheet" href="/static/estilo.css">
        </head>

        <body>

        <h1>Iniciar sesión</h1>

        <form method='POST'>

        Usuario:
        <input type='text' name='usuario'>

        <br><br>

        Contraseña:
        <input type='password' name='password'>

        <br><br>

        <button type='submit'>
        Ingresar
        </button>

        </form>

        </body>
        </html>
    """
@app.route("/logout")
def logout():

    session.pop("usuario", None)

    return """
    <html>

    <head>
    <link rel="stylesheet" href="/static/estilo.css">
    </head>

    <body>

    <img src="/static/Logo LR.jpeg" class="logo">

    <h1>Sesión cerrada</h1>

    <h2>Hasta pronto</h2>

    <p>La sesión se cerró correctamente.</p>

    <a href="/login">
        <button>🔑 Iniciar sesión nuevamente</button>
    </a>

    </body>

    </html>
    """
def crear_base_datos():
    conexion = sqlite3.connect("inventario.db")

    cursor = conexion.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        marca TEXT,
        cantidad INTEGER,
        costo REAL,
        precio REAL
    )
    """)

    conexion.commit()
    conexion.close()
@app.route("/eliminar/<int:id>")
def eliminar(id):

    conexion = sqlite3.connect("inventario.db")
    cursor = conexion.cursor()

    cursor.execute(
        "DELETE FROM productos WHERE id = ?",
        (id,)
    )

    conexion.commit()
    conexion.close()

    return """
    <h1>Producto eliminado</h1>
    <a href='/inventario'>
        <button>Volver</button>
    </a>
    """

@app.route("/inventario")
def inventario():

    buscar = request.args.get("buscar")

    conexion = sqlite3.connect("inventario.db")
    cursor = conexion.cursor()

    if buscar:
        cursor.execute(
            "SELECT * FROM productos WHERE nombre LIKE ?",
            ('%' + buscar + '%',)
        )
    else:
        cursor.execute("SELECT * FROM productos")

    productos = cursor.fetchall()

    conexion.close()

    tabla = """
    <html>

    <head>
    <link rel="stylesheet" href="/static/estilo.css">
    </head>

    <body>

    <h1>📦 Inventario L&R</h1>

    <a href="/">
        <button>🏠 Menú principal</button>
    </a>

<br><br>

<form method="GET">

    <div class="buscador">

        Buscar:
        <input type="text" name="buscar">

        <button type="submit">
            🔍 Buscar
        </button>

    </div>

</form>

    <br>

    <table>
<tr>
    <th>ID</th>
    <th>Nombre</th>
    <th>Marca</th>
    <th>Cantidad</th>
    <th>Costo</th>
    <th>Precio</th>
    <th>Acciones</th>
</tr>
"""

    for producto in productos:
        tabla += f"""
<tr>
    <td>{producto[0]}</td>
    <td>{producto[1]}</td>
    <td>{producto[2]}</td>
    <td>{producto[3]}</td>
    <td>{producto[4]}</td>
    <td>{producto[5]}</td>
    <td>
    <a href="/eliminar/{producto[0]}">
        <button class="rosa">Eliminar</button>
    </a>
    <a href="/exportar_excel">
        <button>📊 Exportar Excel</button>
    </a>
    <a href="/restaurar_excel">
        <button>📥 Restaurar Excel</button>
    </a>
</td>
</tr>
"""

    tabla += """
    </body>
    </html>
    """

    return tabla
@app.route("/agregar", methods=["GET","POST"])
def agregar():
    if request.method == "POST":

        nombre = request.form["nombre"]
        marca = request.form["marca"]
        cantidad = request.form["cantidad"]
        costo = request.form["costo"]
        precio = request.form["precio"]

        conexion = sqlite3.connect("inventario.db")
        cursor = conexion.cursor()

        cursor.execute("""
            INSERT INTO productos
            (nombre, marca, cantidad, costo, precio)
            VALUES (?, ?, ?, ?, ?)
        """, (nombre, marca, cantidad, costo, precio))

        conexion.commit()
        conexion.close()

        from flask import redirect
        return redirect("/inventario")
    return """
        <html>

        <head>
        <link rel="stylesheet" href="/static/estilo.css">
        </head>

        <body>

    <h1>📦 Agregar producto</h1>

<a href="/">
    <button>🏠 Menú principal</button>
</a>

<br><br>

<div class="formulario">
<form method="POST">

    Nombre:
    <input type="text" name="nombre">

    <br><br>

    Marca:
    <input type="text" name="marca">

    <br><br>

    Cantidad:
    <input type="number" name="cantidad">

    <br><br>

    Costo:
    <input type="number" step="0.01" name="costo">

    <br><br>

    Precio:
    <input type="number" step="0.01" name="precio">

    <br><br>

    <button type="submit">
        ➕ Guardar producto
    </button>

</form>
</div>
</body>
</html>
    """
crear_base_datos()
@app.route("/entrada", methods=["GET","POST"])
def entrada():

    if request.method == "POST":

        id_producto = request.form["id_producto"]
        cantidad = request.form["cantidad"]

        conexion = sqlite3.connect("inventario.db")
        cursor = conexion.cursor()

        cursor.execute(
        """
        UPDATE productos
        SET cantidad = cantidad + ?
        WHERE id = ?
        """,
     (cantidad, id_producto)
        )

        conexion.commit()
        conexion.close()     
    return """
        <html>

        <head>
        <link rel="stylesheet" href="/static/estilo.css">
        </head>

        <body>

        <h1>⬆️ Producto resurtido</h1>

        <a href="/">
            <button>🏠 Menú principal</button>
        </a>

        <br><br>

        <div class="formulario">

        <form method="POST">

        Producto (ID):
        <input type="number" name="id_producto">

        <br><br>

        Cantidad:
        <input type="number" name="cantidad">

        <br><br>

        <button type="submit">
            ⬆️ Producto resurtido
        </button>

        </form>

        </div>

        </body>
        </html>
        <h1>Entrada registrada</h1>
         <a href='/inventario'>
        <button>Volver al inventario</button>
         </a>
         """

    conexion = sqlite3.connect("inventario.db")
    cursor = conexion.cursor()

    cursor.execute("SELECT id, nombre FROM productos")

    productos = cursor.fetchall()

    conexion.close()

    pagina = """
    <h1>Producto resurtido</h1>

    <form method="POST">

        Producto:
        <select name="id_producto">
    """

    for producto in productos:

        pagina += f"""
        <option value="{producto[0]}">
            {producto[1]}
        </option>
        """

    pagina += """
        </select>

        <br><br>

        Cantidad:
        <input type="number" name="cantidad">

        <br><br>

        <button type="submit">
            Registrar
        </button>

    </form>
    """

    return pagina
@app.route("/salida", methods=["GET","POST"])
def salida():
    print(request.method)
    if request.method == "POST":

        nombre = request.form["nombre"]
        cantidad = request.form["cantidad"]

        conexion = sqlite3.connect("inventario.db")
        cursor = conexion.cursor()
        cursor.execute(
        "SELECT id FROM productos WHERE nombre = ?",
        (nombre,)
        )

        resultado = cursor.fetchone()
        if resultado:
            id_producto = resultado[0]
        else:
            conexion.close()
            return "<h1>Producto no encontrado</h1>"

        cursor.execute(
        """
        UPDATE productos
        SET cantidad = cantidad - ?
        WHERE id = ?
        """,
     (cantidad, id_producto)
        )

        conexion.commit()
        conexion.close()     
        return """
        <html>

        <head>
        <link rel="stylesheet" href="/static/estilo.css">
        </head>

        <body>

        <h1>✅ Salida registrada</h1>

        <a href="/inventario">
            <button>📦 Volver al inventario</button>
        </a>

        </body>
        </html>
        """

    conexion = sqlite3.connect("inventario.db")
    cursor = conexion.cursor()

    cursor.execute("SELECT id, nombre FROM productos")

    productos = cursor.fetchall()

    conexion.close()

    pagina = """
    <html>

    <head>
    <link rel="stylesheet" href="/static/estilo.css">
    </head>

    <body>

    <h1>⬇️ Ventas</h1>

    <a href="/">
        <button>🏠 Menú principal</button>
    </a>

    <br><br>

    <div class="formulario">

    <form method="POST">

        Producto:
    <input list="productos" name="nombre">
    <datalist id="productos">
    """
    for producto in productos:
        pagina +=f"""
        <option value="{producto[1]}">
        """



    pagina += """
    </datalist>

    <br><br>

    Cantidad:
    <input type="number" name="cantidad">

    <br><br>

    <button type="submit">
        ⬇️ Ventas
    </button>

    </form>

    </div>

    </body>
    </html>
    """

    return pagina

@app.route("/exportar_excel")
def exportar_excel():

    conexion = sqlite3.connect("inventario.db")
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id, nombre, marca, cantidad, costo, precio
        FROM productos
    """)

    productos = cursor.fetchall()

    conexion.close()

    wb = Workbook()
    ws = wb.active

    ws.title = "Inventario"

    ws.append([
        "ID",
        "Nombre",
        "Marca",
        "Cantidad",
        "Costo",
        "Precio"
    ])

    for producto in productos:
        ws.append(producto)

    archivo = "inventario_respaldo.xlsx"

    wb.save(archivo)

    return send_file(
        archivo,
        as_attachment=True
    )

@app.route("/restaurar_excel", methods=["GET", "POST"])
def restaurar_excel():

    if request.method == "POST":

        archivo = request.files["archivo"]

        wb = load_workbook(archivo)
        ws = wb.active

        conexion = sqlite3.connect("inventario.db")
        cursor = conexion.cursor()

        cursor.execute("DELETE FROM productos")

        for fila in ws.iter_rows(min_row=2, values_only=True):

            id_producto, nombre, marca, cantidad, costo, precio = fila

            cursor.execute("""
                INSERT INTO productos
                (id, nombre, marca, cantidad, costo, precio)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                id_producto,
                nombre,
                marca,
                cantidad,
                costo,
                precio
            ))

        conexion.commit()
        conexion.close()

        return redirect("/inventario")

    return """
    <html>
    <head>
        <title>Restaurar Inventario</title>
    </head>
    <body>

        <h2>📥 Restaurar Inventario</h2>

        <form method="POST" enctype="multipart/form-data">

            <input type="file" name="archivo" required>

            <br><br>

            <button type="submit">
                Restaurar
            </button>

        </form>

    </body>
    </html>
    """
if __name__ == "__main__":
    app.run(debug=True)